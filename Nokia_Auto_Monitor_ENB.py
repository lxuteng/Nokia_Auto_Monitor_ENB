# -*- coding: utf-8 -*-

import pandas
from pandas.core.frame import DataFrame
import openpyxl
import os
import sys
import copy
import multiprocessing
import cx_Oracle
import traceback
import datetime
import sqlite3
import pyDes
import paramiko
import ftplib
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor
import threadpool
import lzma
import stat
# import random
import csv
import time
from io import BytesIO
import base64
from matplotlib import pyplot as plt

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication


def copy_right():
    print('\n')
    print(u"""
       --------------------------------------------------
     -   Welcome to use tools!                            -
    -    Author : xuteng.lin                               -
     -   E_mail : xuteng.lin@nokia-sbell-huanuo.com       -
       --------------------------------------------------
    """)
    print('\n')
    auth_time = int(time.strftime('%Y%m%d', time.localtime(time.time())))

    self_key = 'lxtnokia'
    self_iv = 'nokialxt'
    main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]
    try:
        temp_license = open(os.path.join(main_path, 'license')).read()
    except:
        print(u'>>> 无license文件，请申请!')
        os.system("pause")
        sys.exit()
    k = pyDes.des(self_key,
                  mode=pyDes.CBC,
                  IV=self_iv,
                  pad=None,
                  padmode=pyDes.PAD_PKCS5)
    decryptstr = str(k.decrypt(base64.b64decode(temp_license)),
                     encoding='utf-8').split('-')
    if decryptstr[3] == 'Parameter_checker':
        if auth_time > int(decryptstr[2]):
            print(u'>>> 试用版本已过期，请更新！')
            os.system("pause")
            sys.exit()
    else:
        print(u'>>> license错误，请申请！')
        os.system("pause")
        sys.exit()

    print(u'''
    update log:
    2019-03-03 初版；
    2019-03-09 完成全部基本功能；
    ''')
    print(u'-' * 36)
    print(u'      >>>   starting   <<<')
    print(u'-' * 36)
    print(u'\n')
    time.sleep(1)


class Main:

    def __init__(self):
        """初始化"""
        # copy_right()
        self.main_path = os.path.split(os.path.abspath(sys.argv[0]))[0]
        # 获取配置文件
        self.config_list = {}
        self.get_config()
        self.temp_time_now = datetime.datetime.now()
        self.temp_time = self.temp_time_now.strftime('%Y_%m_%d_%H_%M_%S')
        # self.temp_time = '2019_03_05_21_48_03'

        self.para_list_count = {}

        self.now_pm_time_list = []
        for temp_i in range(self.config_list['config']['获取最近时段数']):
            now_pm_time = datetime.datetime.now()
            now_pm_time = now_pm_time-datetime.timedelta(
                hours=8+(15*temp_i+now_pm_time.minute % 15)/60)
            now_pm_time = datetime.datetime.strftime(now_pm_time, "%Y%m%d.%H%M")
            self.now_pm_time_list.append(now_pm_time)

    def get_time_item(self, time_type, time_num):
        pass

    # 进度条
    @staticmethod
    def progress(num_total, num_run, file_name=''):
        bar_len = 10
        hashes = '|' * int(num_run / num_total * bar_len)
        spaces = '_' * (bar_len - len(hashes))
        sys.stdout.write(
            "\r%s %s %d%%  %s" % (str(num_run), hashes + spaces, int(
                num_run / num_total * 100),
                                  file_name)
        )
        sys.stdout.flush()

    def get_config(self):
        # 初始化配置列表

        path_base_data = os.path.join(
            self.main_path,
            '_config',
            'config.xlsx')
        f_base_data_wb = openpyxl.load_workbook(path_base_data, read_only=True)
        for temp_sheet_name in f_base_data_wb.sheetnames:
            if temp_sheet_name not in self.config_list:
                self.config_list[temp_sheet_name] = {}
                temp_f_base_data_wb_sheet = f_base_data_wb[temp_sheet_name]
                if temp_sheet_name == 'IP':
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        self.config_list[
                            temp_sheet_name][temp_value[0]] = temp_value
                elif temp_sheet_name == 'counter':
                    for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                        temp_value = [j.value for j in temp_row]
                        if temp_value[0] == '多进程':
                            self.config_list[
                                temp_sheet_name
                            ][temp_value[0]] = temp_value[1]
                        else:
                            self.config_list[
                                temp_sheet_name
                            ][temp_value[0]] = temp_value[1].split(',')
                elif temp_sheet_name in ['counter_sql', 'kpi_sql']:
                    temp_iter_rows = temp_f_base_data_wb_sheet.iter_rows()
                    next(temp_iter_rows)
                    for temp_row in temp_iter_rows:
                        temp_value = [j.value for j in temp_row]
                        if temp_value[1] == '启用':
                            self.config_list[
                                temp_sheet_name][temp_value[0]] = temp_value
                else:
                    for temp_row in temp_f_base_data_wb_sheet.iter_rows():
                        temp_value = [j.value for j in temp_row]
                        if temp_value[0] == '通用远程路径':
                            self.config_list[
                                temp_sheet_name
                            ][temp_value[0]] = temp_value[1:]
                        else:
                            self.config_list[
                                temp_sheet_name
                            ][temp_value[0]] = temp_value[1]

    def ftp_process1(self):
        # 多进程到不同的服务器上获取数据
        try:
            processes_num = self.config_list['config']['多进程']
            process_pool = multiprocessing.Pool(processes=processes_num)
            for temp_db in self.config_list['IP']:
                process_pool.apply_async(
                    self.ftp_operator,
                    args=(self.config_list['IP'][temp_db]
                          )
                )
            process_pool.close()
            process_pool.join()
        except:
            traceback.print_exc()

    def ftp_process(self):
        processes_num = self.config_list['config']['多进程']
        pool = threadpool.ThreadPool(processes_num)
        name_list = [
            self.config_list['IP'][temp_db] for temp_db in self.config_list[
                'IP']]
        requests = threadpool.makeRequests(self.ftp_operator, name_list)
        [pool.putRequest(req) for req in requests]
        pool.wait()

    def ftp_operator(self, ip_config):
        temp_ip = ip_config[0]
        if ip_config[1] is not None:
            temp_usr = ip_config[1]
        else:
            temp_usr = self.config_list['config']['通用账号']

        if ip_config[2] is not None:
            temp_pwr = ip_config[2]
        else:
            temp_pwr = self.config_list['config']['通用密码']

        if ip_config[3] is not None:
            temp_remote_path = [ip_config[3],]
        else:
            temp_remote_path = self.config_list['config']['通用远程路径']

        temp_local_path = self.config_list['config']['PM文件本地保存路径']

        status = 0
        sftp = []
        try:
            try:
                T = paramiko.Transport(temp_ip)
                T.connect(username=temp_usr, password=temp_pwr)
                sftp = [paramiko.SFTPClient.from_transport(T), 'sftp']
                print('>>> SFTP connect successful!\n')
                status = 1
            except:
                sftp = [ftplib.FTP(temp_ip,temp_usr,temp_pwr,timeout=3)
                    ,'ftp']
                print('>>> FTP connect successful!\n')
                status = 1
        except:
            print('>>> fail connect to:', temp_ip, '\n')

        try:
            if status == 1:
                if sftp[1] == 'sftp':
                    temp_iii = 0
                    for temp_remote in temp_remote_path:
                        try:
                            for i in sftp[0].listdir_attr(temp_remote):
                                temp_filename = i.filename
                                if 'PM.BTS-' in temp_filename and temp_filename[
                                                    -7:] == '.xml.xz':
                                    for temp_time in self.now_pm_time_list:
                                        if temp_time in temp_filename:
                                            full_i = '/'.join((temp_remote,
                                                               temp_filename))
                                            new_full_i = os.path.join(
                                                temp_local_path, temp_filename)
                                            sftp[0].get(full_i, new_full_i)
                                            temp_iii = 1
                        except:
                            pass
                            # traceback.print_exc()
                        if temp_iii == 1:
                            break
        except:
            traceback.print_exc()

    def get_files(self):
        print('>>> 获取本地数据...')
        self.file_list = []

        for file in os.listdir(self.config_list['config']['PM文件本地保存路径']):
            if 'PM.BTS-' in file and '.xml.xz' in file:
                self.file_list.append(os.path.join(
                    self.config_list['config']['PM文件本地保存路径'],
                    file))
        file_n = len(self.file_list)
        if file_n == 0:
            print('>>> 未获取到原始文件，请检查！')
            sys.exit()
        else:
            print('>>> 获取原始文件：', file_n)

    def parser(self, temp_file):
        if os.path.split(temp_file)[-1].split('.')[-1] == 'xz':
            try:
                tree = ET.parse(lzma.open(temp_file, 'rb'))
            except:
                traceback.print_exc()
        elif os.path.split(temp_file)[-1].split('.')[-1] == 'xml':
            tree = ET.parse(temp_file)

        temp_data = {}
        temp_data_format = []
        root = tree.getroot()

        # 获取时间
        for temp_time in root:
            temp_pm_time = temp_time.attrib['startTime'].split('+')[0].split('.')[0]
            temp_pm_time = datetime.datetime.strptime(temp_pm_time, "%Y-%m-%dT%H:%M:%S")
            pm_time = temp_pm_time + datetime.timedelta(hours=8)
            pm_time = datetime.datetime.strftime(pm_time, "%Y%m%d%H%M")
            if pm_time not in temp_data:
                temp_data[pm_time] = {}
            break
        for temp_iter in root.iter(tag='PMMOResult'):
            for i in temp_iter.iter(tag='localMoid'):
                try:
                    temp_dn = '{0}_{1}'.format(i.text.split('-')[2][:6], i.text.split('-')[3].split('/')[0])
                except:
                    pass
            for j in temp_iter.iter(tag='NE-WBTS_1.0'):
                if j.attrib['measurementType'] in self.config_list[
                        'counter']['pm_list']:
                    if temp_dn not in temp_data[pm_time]:
                        temp_data[pm_time][temp_dn] = {}
                    for k in j:
                        if k.tag in self.config_list['counter']['counter_list']:
                            temp_data[pm_time][temp_dn][k.tag] = k.text
        for temp_time in temp_data:
            for temp_cellid in temp_data[temp_time]:
                temp_value = [temp_time, temp_cellid[:6], temp_cellid]
                for temp_counter in self.config_list['counter']['counter_list']:
                    if temp_counter in temp_data[temp_time][temp_cellid]:
                        temp_value.append(
                            temp_data[temp_time][temp_cellid][temp_counter])
                    else:
                        temp_value.append('None')
                temp_data_format.append(temp_value)
        return temp_data_format

    @staticmethod
    def all_parser_value_gather(temp_value):
        global all_parser_value_list
        try:
            all_parser_value_list += temp_value
        except:
            traceback.print_exc()

    def circuit(self):
        print('>>> 开始解码...')

        global all_parser_value_list
        all_parser_value_list = []

        progress_n = 0
        file_n = len(self.file_list)
        processes_num = self.config_list['counter']['多进程']
        process_pool = multiprocessing.Pool(processes=processes_num)
        try:
            for temp_file in self.file_list:
                process_pool.apply_async(
                    self.parser,
                    args=(temp_file,),
                    callback=self.all_parser_value_gather
                )

            process_pool.close()
            process_pool.join()
        except:
            traceback.print_exc()

        self.progress(file_n, progress_n, 'parse finish!\n')

        f_csv_parser = os.path.join(
            self.config_list['config']['PM文件解析结果保存路径'],
            ''.join((
                'kpi_parser_', self.temp_time, '.csv'
            ))
        )
        f = open(f_csv_parser, 'w', newline='')
        f_csv = csv.writer(f)
        f_csv.writerow(
            ['SDATE', 'ENBID', 'ENB_CELLID'] + self.config_list[
                'counter']['counter_list']
        )
        f_csv.writerows(all_parser_value_list)
        print('>>> 解码完成!')

    def online_db_input_warehousing_temp(self):

        f_csv_parser = os.path.join(
            self.config_list['config']['PM文件解析结果保存路径'],
            ''.join((
                'kpi_parser_', self.temp_time, '.csv'
            ))
        )

        f_csv = pandas.read_csv(f_csv_parser)

        try:
            # local_conn = sqlite3.connect(
            #     os.path.join(
            #         self.config_list['config']['数据库'], 'db_counter.db'
            #     ),
            #     check_same_thread=False
            # )
            #
            local_conn = sqlite3.connect(
                ':memory:',
                check_same_thread=False
            )

            print('>>> counter:', f_csv_parser, ':入库中...')
            f_csv.to_sql(
                'kpi_list',
                con=local_conn,
                if_exists='append',
                chunksize=500
            )
            print('>>> counter:', f_csv_parser, ':入库完毕...')
            return local_conn
            # local_conn.close()
        except:
            traceback.print_exc()

    def local_db_operator_temp(self, local_conn):

        # local_conn = sqlite3.connect(
        #     os.path.join(
        #         self.config_list['config']['数据库'], 'db_counter.db'
        #     ),
        #     check_same_thread=False
        # )

        cu = local_conn.cursor()
        for temp_local_sql in self.config_list['counter_sql']:
            print('='*32)
            print('>>>获取本地数据', temp_local_sql)
            f_sql = open(
                os.path.join(
                    self.main_path,
                    '_sql/counter_sql',
                    ''.join((temp_local_sql, '.sql'))
                ), encoding='gbk'
            )
            sql_scr = f_sql.read()
            cu.execute(sql_scr)

            kpi_list = os.path.join(
                self.config_list['config']['PM文件解析结果保存路径'],
                ''.join((
                    temp_local_sql, '_', self.temp_time, '.csv'
                ))
            )
            f = open(kpi_list, 'w', newline='')
            f_csv = csv.writer(f)

            temp_head = [i[0] for i in cu.description]
            temp_value_list = cu.fetchall()
            f_csv.writerow(temp_head)
            f_csv.writerows(temp_value_list)

    def online_db_input_warehousing(self):

        f_csv_parser = os.path.join(
            self.config_list['config']['PM文件解析结果保存路径'],
            ''.join((
                'cell_day_all_', self.temp_time, '.csv'
            ))
        )

        f_csv = pandas.read_csv(f_csv_parser, encoding='gbk')

        try:
            local_conn = sqlite3.connect(
                os.path.join(
                    self.config_list['config']['数据库'], 'db.db'
                ),
                check_same_thread=False
            )

            print('>>> kpi:', f_csv_parser, ':入库中...')
            f_csv.to_sql(
                'kpi_list',
                con=local_conn,
                if_exists='append',
                chunksize=500
            )
            print('>>> cell_day_all :', f_csv_parser, ':入库完毕...')
            return local_conn
            # local_conn.close()
        except:
            traceback.print_exc()

    def local_db_operator(self, local_conn):

        # local_conn = sqlite3.connect(
        #     os.path.join(
        #         self.config_list['config']['数据库'], 'db_counter.db'
        #     ),
        #     check_same_thread=False
        # )

        cu = local_conn.cursor()
        for temp_local_sql in self.config_list['kpi_sql']:
            print('='*32)
            print('>>>获取本地数据', temp_local_sql)
            f_sql = open(
                os.path.join(
                    self.main_path,
                    '_sql/kpi_sql',
                    ''.join((temp_local_sql, '.sql'))
                ), encoding='gbk'
            )
            sql_scr = f_sql.read()
            cu.execute(sql_scr)

            kpi_list = os.path.join(
                self.config_list['config']['PM文件解析结果保存路径'],
                ''.join((
                    temp_local_sql, '_', self.temp_time, '.csv'
                ))
            )
            f = open(kpi_list, 'w', newline='')
            f_csv = csv.writer(f)

            temp_head = [i[0] for i in cu.description]
            temp_value_list = cu.fetchall()
            f_csv.writerow(temp_head)
            f_csv.writerows(temp_value_list)


if __name__ == '__main__':
    multiprocessing.freeze_support()
    print(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
    star_time = time.time()
    main = Main()
    print(main.config_list['counter'])
    # main.ftp_process()
    main.get_files()
    main.circuit()
    main.local_db_operator_temp(main.online_db_input_warehousing_temp())
    main.local_db_operator(main.online_db_input_warehousing())

    print(''.join((time.strftime('%Y/%m/%d %H:%M:%S', time.localtime()))))
    print(''.join(('>>> 历时：', time.strftime(
        '%Y/%m/%d %H:%M:%S',
        time.gmtime(time.time() - star_time)
        )
    )))

